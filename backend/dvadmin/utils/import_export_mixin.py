# -*- coding: utf-8 -*-
from urllib.parse import quote

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework.decorators import action
from rest_framework.request import Request

from dvadmin.utils.import_export import import_to_data
from dvadmin.utils.json_response import DetailResponse
from dvadmin.utils.request_util import get_verbose_name

#注释编号: django-vue3-admin_import_export_mixin490812: 导出需求，主要是针对归属部门进行特殊处理
from dvadmin.system.models import Dept


import pandas as pd
import codecs
import numpy as np 




class ImportSerializerMixin:
    """
    自定义导入模板、导入功能
    """

    # 导入字段
    import_field_dict = {}
    # 导入序列化器
    import_serializer_class = None
    # 表格表头最大宽度，默认50个字符
    export_column_width = 50

    def is_number(self,num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        获取字符串最大长度
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return round(length, 1) if length <= self.export_column_width else self.export_column_width

    @action(methods=['get','post'],detail=False)
    @transaction.atomic  # Django 事务,防止出错
    def import_data(self, request: Request, *args, **kwargs):
        """
        导入模板
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        assert self.import_field_dict, "'%s' 请配置对应的导出模板字段。" % self.__class__.__name__
        # 导出模板
        if request.method == "GET":
            # 示例数据
            queryset = self.filter_queryset(self.get_queryset())
            # 导出excel 表
            response = HttpResponse(content_type="application/msexcel")
            response["Access-Control-Expose-Headers"] = f"Content-Disposition"
            response[
                "Content-Disposition"
            ] = f'attachment;filename={quote(str(f"导入{get_verbose_name(queryset)}模板.xlsx"))}'
            wb = Workbook()
            ws1 = wb.create_sheet("data", 1)
            ws1.sheet_state = "hidden"
            ws = wb.active
            row = get_column_letter(len(self.import_field_dict) + 1)
            column = 10
            header_data = [
                "序号",
            ]
            validation_data_dict = {}
            for index, ele in enumerate(self.import_field_dict.values()):
                if isinstance(ele, dict):
                    header_data.append(ele.get("title"))
                    choices = ele.get("choices", {})
                    if choices.get("data"):
                        data_list = []
                        data_list.extend(choices.get("data").keys())
                        validation_data_dict[ele.get("title")] = data_list
                    elif choices.get("queryset") and choices.get("values_name"):
                        data_list = choices.get("queryset").values_list(choices.get("values_name"), flat=True)
                        validation_data_dict[ele.get("title")] = list(data_list)
                    else:
                        continue
                    column_letter = get_column_letter(len(validation_data_dict))
                    dv = DataValidation(
                        type="list",
                        formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[ele.get('title')]) + 1}",
                        allow_blank=True,
                    )
                    ws.add_data_validation(dv)
                    dv.add(f"{get_column_letter(index + 2)}2:{get_column_letter(index + 2)}1048576")
                else:
                    header_data.append(ele)
            # 添加数据列
            ws1.append(list(validation_data_dict.keys()))
            for index, validation_data in enumerate(validation_data_dict.values()):
                for inx, ele in enumerate(validation_data):
                    ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
            # 插入导出模板正式数据
            df_len_max = [self.get_string_len(ele) for ele in header_data]
            ws.append(header_data)
            # 　更新列宽
            for index, width in enumerate(df_len_max):
                ws.column_dimensions[get_column_letter(index + 1)].width = width
            tab = Table(displayName="Table1", ref=f"A1:{row}{column}")  # 名称管理器
            style = TableStyleInfo(
                name="TableStyleLight11",
                showFirstColumn=True,
                showLastColumn=True,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            wb.save(response)
            return response
        else:
            # 从excel中组织对应的数据结构，然后使用序列化器保存
            queryset = self.filter_queryset(self.get_queryset())
            # 获取多对多字段
            m2m_fields = [
                ele.name
                for ele in queryset.model._meta.get_fields()
                if hasattr(ele, "many_to_many") and ele.many_to_many == True
            ]
            import_field_dict = {'id':'更新主键(勿改)',**self.import_field_dict}

            # 注释编号:django-vue3-admin__import_export_mixin550918:代码开始行
            # 功能说明:这里增加了一个serializer_fields字段传进函数，主要是为了后面再验证字段类型使用的
            serializer_fields = self.get_serializer().get_fields()
            data = import_to_data(request.data.get("url"), import_field_dict, serializer_fields, m2m_fields)
            # 注释编号:django-vue3-admin__import_export_mixin550918:代码结束行
            
            for ele in data:
                filter_dic = {'id':ele.get('id')}
                instance = filter_dic and queryset.filter(**filter_dic).first()
                # print(156,ele)
                serializer = self.import_serializer_class(instance, data=ele, request=request)
                serializer.is_valid(raise_exception=True)
                serializer.save()
            return DetailResponse(msg=f"导入成功！")

    @action(methods=['get'],detail=False)
    def update_template(self,request):
        queryset = self.filter_queryset(self.get_queryset())
        assert self.import_field_dict, "'%s' 请配置对应的导入模板字段。" % self.__class__.__name__
        assert self.import_serializer_class, "'%s' 请配置对应的导入序列化器。" % self.__class__.__name__
        data = self.import_serializer_class(queryset, many=True, request=request).data
        # 导出excel 表
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"导出{get_verbose_name(queryset)}.xlsx"))}'
        wb = Workbook()
        ws1 = wb.create_sheet("data", 1)
        ws1.sheet_state = "hidden"
        ws = wb.active
        import_field_dict = {}
        header_data = ["序号","更新主键(勿改)"]
        hidden_header = ["#","id"]
        #----设置选项----
        validation_data_dict = {}
        for index, item in enumerate(self.import_field_dict.items()):
            items = list(item)
            key = items[0]
            value = items[1]
            if isinstance(value, dict):
                header_data.append(value.get("title"))
                hidden_header.append(value.get('display'))
                choices = value.get("choices", {})
                if choices.get("data"):
                    data_list = []
                    data_list.extend(choices.get("data").keys())
                    validation_data_dict[value.get("title")] = data_list
                elif choices.get("queryset") and choices.get("values_name"):
                    data_list = choices.get("queryset").values_list(choices.get("values_name"), flat=True)
                    validation_data_dict[value.get("title")] = list(data_list)
                else:
                    continue
                column_letter = get_column_letter(len(validation_data_dict))
                dv = DataValidation(
                    type="list",
                    formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[value.get('title')]) + 1}",
                    allow_blank=True,
                )
                ws.add_data_validation(dv)
                dv.add(f"{get_column_letter(index + 3)}2:{get_column_letter(index + 3)}1048576")
            else:
                header_data.append(value)
                hidden_header.append(key)
        # 添加数据列
        ws1.append(list(validation_data_dict.keys()))
        for index, validation_data in enumerate(validation_data_dict.values()):
            for inx, ele in enumerate(validation_data):
                ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
        #--------
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(hidden_header) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key, val in results.items():
                    if key == h_item:
                        if val is None or val == "":
                            results_list.append("")
                        elif isinstance(val,list):
                            results_list.append(str(val))
                        else:
                            results_list.append(val)
                        # 计算最大列宽度
                        if isinstance(val,str):
                            result_column_width = self.get_string_len(val)
                            if h_index != 0 and result_column_width > df_len_max[h_index]:
                                df_len_max[h_index] = result_column_width
            ws.append([index+1,*results_list])
            column += 1
        # 　更新列宽
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # 名称管理器
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(response)
        return response


class ExportSerializerMixin:
    """
    自定义导出功能
    """

    # 导出字段
    export_field_label = []
    # 导出序列化器
    export_serializer_class = None
    # 表格表头最大宽度，默认50个字符
    export_column_width = 50

    def is_number(self,num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        获取字符串最大长度
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return round(length, 1) if length <= self.export_column_width else self.export_column_width


    # 注释编号:django-vue3-admin_import_export_mixin564615
    # 功能说明:这一块是使用pandas进行配置导出的功能

    @action(methods=['get'],detail=False)
    def export_data(self, request: Request, *args, **kwargs):
        """
        导出功能
        :param request:
        :param args:
        :param kwargs:
        :return:
        """      
        queryset = self.filter_queryset(self.get_queryset())
        # assert self.export_field_label, "'%s' 请配置对应的导出模板字段。" % self.__class__.__name__   

        # 配置导出csv,不要导出excel表,因为excel表导出太慢了,容易超时
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"

        
        cur_model = queryset.model   # 拿到当前queryset对应的model



        # 注释编号:django-vue3-admin_import_export_mixin254314
        # 功能说明:在这里判断是否在传进来export_field_label，如果有，就按export_field_label的需求导出，如果没有，就自己拼接，导出所有数据
        #这里主要就是对于导出来列名进行处理，有外键的折拼接上_id后缀，而没有的，直接放在export_field_label里面，并且把对应的中文名称verbose_name也加上字段内

        if self.export_field_label:
            key_list = list(self.export_field_label.keys()) 
        else:
            #这里就要自己拼接出来一个字典，字典格式如下
            # export_field_label = {"title": "区域",}
            export_field_label = {}
            for field  in cur_model._meta.fields:
                field_name = None
                if field.many_to_one:
                    field_name = field.name + '_id'   #这里还要组装一下_id的后缀
                    export_field_label[field_name] = field.verbose_name
                else:
                    field_name = field.name   #这里就不需要再拼接_id
                    export_field_label[field_name] = field.verbose_name

            key_list = list(export_field_label.keys())
        # 注释编号:django-vue3-admin_import_export_mixin254314

        
        
        try:
            #将queryset数据转为pandas对应的DataFrame数据,而且只取相应的列数据出来columns=key_list
            #这个df数据就是已经序列好的数据
            df = pd.DataFrame(queryset.values(), columns=key_list)
        except:
            pass


        # 注释编号:django-vue3-admin_import_export_mixin285910
        # 功能说明:这里是对于导出来列内容的处理，主要是处理那些外健导出来是要转为对应的外健的值，而不是直接导出来外健的ID
        

        for key in key_list:
            #这里是判断export_field_label中的key是否传有带_id的字段过来,如果有,就说明想导出相应的外健对应的值
            # 如果两个值不等,那说明要是导出外健对应的值的
            newKey =  key.replace('_id', '') 

            if newKey != key:
                try:
                    #这里是拿到当前cur_model对应的newKey的外健对象的所有值并转为这字典

                    if key == "dept_belong_id":   # 这里专门对归属部门一项进行特殊处理
                        Dept_queryset = Dept.objects.all()
                        choices_dict = pd.DataFrame(Dept_queryset.values(), columns=["id", "name"])   #只拿其中的id与name字段便可，其它的不需要了。
                        choices_dict = dict(choices_dict.values.tolist()) #要先拿到df的values先转为tolist，再全部转为dict便可
                        df[f'{key}'] = pd.to_numeric(df[f'{key}'], errors='coerce')   # 将字符串转为数字num
                    else:
                        #注释编号: django-vue3-admin-import_export_mixin031018:这里如果相应的newkey对应的外键返回不是str类型，那就无法正常导出，解决方法就是在model里面配置__str__方法return一个string类型的值
                        choices_dict = dict(getattr(cur_model, newKey).field.get_choices())
                    df[f'{key}'] =  df[f'{key}'].replace(choices_dict)
                except Exception as e :
                    # 注释编号:django-vue3-admin-import_export_mixin111316
                    # 功能说明:这里status=599要与前端自定义的error_code=599对应上，不然前端无法识别到错误
                    # 功能说明:这里要判断一下是出现什么错误，如果是出现'__str__ returned non-string (type int)'错误，就返回一个status=599的HttpResponse
                    if (str(e) == "__str__ returned non-string (type int)"):
                        return HttpResponse(status=599) 
                    # 注释编号:django-vue3-admin-import_export_mixin111316
                    else:
                        #即是出现其它错误如'NoneType' object has no attribute 'model'不用理会
                        pass

            elif key == 'modifier':   #这里专门对修改人的一项进行特殊处理
                    choices_dict = dict(getattr(cur_model, 'creator').field.get_choices())
                    df[f'{key}'] = pd.to_numeric(df[f'{key}'], errors='coerce')  # 将字符串转为数字num
                    df[f'{key}'] =  df[f'{key}'].replace(choices_dict)


        # 注释编号:django-vue3-admin_import_export_mixin285910


        #列名根据当前export_field_label传过来的字典进行匹配更换,要判断一下是使用传进来的需求，还是自己拼接所有字段
        if self.export_field_label:
            df = df.rename(columns=dict(self.export_field_label))
        else:
            df = df.rename(columns=dict(export_field_label))
       

        # 注释编号:django-vue3-admin_import_export_mixin382615
        # 功能说明:处理可能存在的一些=号被视为公式的问题，如果有=开头数据，直接替换为，但是这样只要=开头的数据就会直接整个数据被替换掉了，这样不合理的
        # def clear_formula(cell):  
        #     if isinstance(cell, str) and cell.startswith("="): 
        #         return np.nan  # 将公式置为NaN，这样在导出时就不会被当作公式了  
        #     return cell  
        # df = df.applymap(clear_formula) 
        # 注释编号:django-vue3-admin_import_export_mixin382615


        # 注释编号:django-vue3-admin_import_export_mixin532815
        # 功能说明:如果在导出csv文档就启用这里
        response["content-disposition"] = f'attachment;filename={quote(str(f"导出{get_verbose_name(queryset)}.csv"))}' 
        response.write(codecs.BOM_UTF8)   # 这里是为了防止写入csv文件时出现乱码
        df.to_csv(response, index=False)  # 这个是最终处理好的数据传给response对象，最终返回给前端下载，index=False是说明不添加
        # 注释编号:django-vue3-admin_import_export_mixin532815

        
        # 注释编号:django-vue3-admin_import_export_mixin372915
        # 功能说明:如果是导xlsx文档就启用这里
        # response["content-disposition"] = f'attachment;filename={quote(str(f"导出{get_verbose_name(queryset)}.xlsx"))}'
        # df.to_excel(response, index=False)
        # 注释编号:django-vue3-admin_import_export_mixin372915



        return response
        # 把处理好的response返回去给前端
        # ---------------这个函数是制作导出csv功能的

         # 注释编号:django-vue3-admin_import_export_mixin564615

